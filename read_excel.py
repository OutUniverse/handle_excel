from openpyxl import *
from openpyxl.styles import Alignment, Font

class HandleExcel():
    def __init__(self, filenname):
        self.filename = filenname
        self.handle = load_workbook(filenname)
        self.active_sheet = self.handle.active

    # Read all values in a column    
    def read_all_col(self, column_name):
        max_row = self.active_sheet.max_row
        column_data = []

        for i in range(1, max_row + 1):
            column_data.append(self.active_sheet[column_name + str(i)].value)

        return column_data
    
    def write_list_to_column(self, data, col_num):
        for i, item in enumerate(data):
            # if i == 0:
            self.active_sheet.cell(column=col_num, row=i+1).alignment = Alignment(horizontal='center', vertical='center')
            self.active_sheet.cell(column=col_num, row=i+1).font  = Font(name='Times New Roman')
            self.active_sheet.cell(column=col_num, row=i+1).value = item

        self.handle.save(self.filename)
 
if __name__ == "__main__":
    read_file = HandleExcel('C:\\Users\\Administrator\\Desktop\\数据处理-zyx\\20230522\\sample1\\1-3\\S1-(1-3)-Total.xlsx')
    read_file.test()
